from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx')

ws = wb.active

#ws['A2'].value = "Esther" 

#wb.save('Grades.xlsx')

# printing sheetnames
#print(wb.sheetnames)
# Creating a new sheet
#wb.create_sheet("Test")
#print(wb.sheetnames)
#wb.save('Grades.xlsx')

#Accessing a sheet
#print(wb['Sheet1'])
